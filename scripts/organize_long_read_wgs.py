#!/usr/bin/env python3

import argparse
import csv
import re
import unicodedata
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from workspace import PROJECT_ROOT, resolve_workspace


ROOT = PROJECT_ROOT
RAW_DIR = ROOT / "raw" / "sources"
SOURCE_PAGES_DIR = ROOT / "wiki" / "sources"
MANIFEST_PATH = ROOT / "long_read_wgs_manifest.tsv"
XLSX_PATH = ROOT / "long_read_wgs_papers.xlsx"
LOG_PATH = ROOT / "wiki" / "log.md"


def configure_workspace(root: Path):
    global ROOT, RAW_DIR, SOURCE_PAGES_DIR, MANIFEST_PATH, XLSX_PATH, LOG_PATH
    ROOT = root
    RAW_DIR = ROOT / "raw" / "sources"
    SOURCE_PAGES_DIR = ROOT / "wiki" / "sources"
    MANIFEST_PATH = ROOT / "long_read_wgs_manifest.tsv"
    XLSX_PATH = ROOT / "long_read_wgs_papers.xlsx"
    LOG_PATH = ROOT / "wiki" / "log.md"

RECORDS = [
    {
        "published_date": "2019-04-16",
        "first_author": "Chaisson",
        "title": "Multi-platform discovery of haplotype-resolved structural variation in human genomes",
        "article_url": "https://www.nature.com/articles/s41467-018-08148-z",
        "pdf_url": "https://www.nature.com/articles/s41467-018-08148-z.pdf",
        "legacy_filename": "2019_chaisson_multi-platform-haplotype-resolved-structural-variation-human-genomes.pdf",
        "focus": "foundational human SV benchmark across multiple long-read-aware platforms",
    },
    {
        "published_date": "2022-09-20",
        "first_author": "Otsuki",
        "title": "Construction of a trio-based structural variation panel utilizing activated T lymphocytes and long-read sequencing technology",
        "article_url": "https://www.nature.com/articles/s42003-022-03953-1",
        "pdf_url": "https://www.nature.com/articles/s42003-022-03953-1.pdf",
        "legacy_filename": "2022_otsuki_trio-structural-variation-panel-activated-t-lymphocytes.pdf",
        "focus": "population-scale trio long-read SV resource",
    },
    {
        "published_date": "2022-10-09",
        "first_author": "Kobayashi",
        "title": "Approaches to long-read sequencing in a clinical setting to improve diagnostic rate",
        "article_url": "https://www.nature.com/articles/s41598-022-20113-x",
        "pdf_url": "https://www.nature.com/articles/s41598-022-20113-x.pdf",
        "legacy_filename": "2022_sanford-kobayashi_clinical-long-read-diagnostic-rate.pdf",
        "focus": "clinical long-read WGS for improved diagnosis",
    },
    {
        "published_date": "2023-08-24",
        "first_author": "Hard",
        "title": "Long-read whole-genome analysis of human single cells",
        "article_url": "https://www.nature.com/articles/s41467-023-40898-3",
        "pdf_url": "https://www.nature.com/articles/s41467-023-40898-3.pdf",
        "legacy_filename": "2023_hard_long-read-whole-genome-analysis-human-single-cells.pdf",
        "focus": "single-cell long-read WGS architecture and somatic variation",
    },
    {
        "published_date": "2024-01-29",
        "first_author": "Mahmoud",
        "title": "Utility of long-read sequencing for All of Us",
        "article_url": "https://www.nature.com/articles/s41467-024-44804-3",
        "pdf_url": "https://www.nature.com/articles/s41467-024-44804-3.pdf",
        "legacy_filename": "2024_mahmoud_utility-long-read-sequencing-all-of-us.pdf",
        "focus": "large-cohort long-read WGS evaluation for medically relevant genes",
    },
    {
        "published_date": "2024-04-17",
        "first_author": "Kosugi",
        "title": "Comparative evaluation of SNVs, indels, and structural variations detected with short- and long-read sequencing data",
        "article_url": "https://www.nature.com/articles/s41439-024-00276-x",
        "pdf_url": "https://www.nature.com/articles/s41439-024-00276-x.pdf",
        "legacy_filename": "2024_kosugi_comparative-evaluation-short-vs-long-read-variants.pdf",
        "focus": "short-read vs long-read WGS comparison across variant classes",
    },
    {
        "published_date": "2024-12-18",
        "first_author": "Showpnil",
        "title": "Long-read genome sequencing resolves complex genomic rearrangements in rare genetic syndromes",
        "article_url": "https://www.nature.com/articles/s41525-024-00454-4",
        "pdf_url": "https://www.nature.com/articles/s41525-024-00454-4.pdf",
        "legacy_filename": "2024_showpnil_long-read-genome-sequencing-complex-rearrangements-rare-syndromes.pdf",
        "focus": "rare disease and complex rearrangement resolution",
    },
    {
        "published_date": "2025-02-10",
        "first_author": "Gong",
        "title": "Long-read sequencing of 945 Han individuals identifies structural variants associated with phenotypic diversity and disease susceptibility",
        "article_url": "https://www.nature.com/articles/s41467-025-56661-9",
        "pdf_url": "https://www.nature.com/articles/s41467-025-56661-9.pdf",
        "legacy_filename": "2025_gong_long-read-sequencing-945-han-individuals.pdf",
        "focus": "population SV discovery and phenotype associations",
    },
    {
        "published_date": "2025-03-14",
        "first_author": "Sinha",
        "title": "Long read sequencing enhances pathogenic and novel variation discovery in patients with rare diseases",
        "article_url": "https://www.nature.com/articles/s41467-025-57695-9",
        "pdf_url": "https://www.nature.com/articles/s41467-025-57695-9.pdf",
        "legacy_filename": "2025_sinha_long-read-sequencing-rare-diseases.pdf",
        "focus": "clinical rare disease discovery including methylation-aware workflow",
    },
    {
        "published_date": "2025-07-23",
        "first_author": "Schloissnig",
        "title": "Structural variation in 1,019 diverse humans based on long-read sequencing",
        "article_url": "https://www.nature.com/articles/s41586-025-09290-7",
        "pdf_url": "https://www.nature.com/articles/s41586-025-09290-7.pdf",
        "legacy_filename": "2025_schloissnig_structural-variation-1019-diverse-humans-long-read-sequencing.pdf",
        "alternate_filenames": [
            "schloissnig_2025_structural_variation_in_1_019.pdf",
        ],
        "focus": "diverse population-scale long-read SV atlas",
    },
]


def ascii_text(value):
    normalized = unicodedata.normalize("NFKD", value)
    return normalized.encode("ascii", "ignore").decode("ascii")


def slug_token(value):
    value = ascii_text(value).lower()
    value = re.sub(r"[^a-z0-9-]+", "_", value)
    value = re.sub(r"_+", "_", value)
    return value.strip("_")


def title_prefix(title):
    normalized_title = re.sub(r"(?<=\d),(?=\d)", "", ascii_text(title))
    words = re.findall(r"[A-Za-z0-9-]+", normalized_title)
    return "_".join(word.lower() for word in words[:5])


def enrich_record(record):
    enriched = dict(record)
    enriched["year"] = record["published_date"][:4]
    enriched["filename"] = "%s_%s_%s.pdf" % (
        slug_token(record["first_author"]),
        enriched["year"],
        title_prefix(record["title"]),
    )
    enriched["local_path"] = "raw/sources/%s" % enriched["filename"]
    enriched["source_page"] = "wiki/sources/%s.md" % Path(record["legacy_filename"]).stem
    return enriched


def rename_pdfs(records):
    renamed = []
    for record in records:
        target_path = RAW_DIR / record["filename"]
        source_candidates = [record["legacy_filename"]]
        source_candidates.extend(record.get("alternate_filenames", []))

        if target_path.exists():
            continue

        for candidate_name in source_candidates:
            candidate_path = RAW_DIR / candidate_name
            if candidate_path.exists():
                candidate_path.rename(target_path)
                renamed.append((candidate_name, record["filename"]))
                break

    return renamed


def update_source_pages(records):
    updated = []
    for record in records:
        page_path = ROOT / record["source_page"]
        if not page_path.exists():
            continue

        new_path = "raw/sources/%s" % record["filename"]
        content = page_path.read_text(encoding="utf-8")
        updated_content = content
        replaced = False
        for candidate_name in [record["legacy_filename"]] + record.get("alternate_filenames", []):
            old_path = "raw/sources/%s" % candidate_name
            if old_path in updated_content:
                updated_content = updated_content.replace(old_path, new_path)
                replaced = True

        if replaced:
            page_path.write_text(updated_content, encoding="utf-8")
            updated.append(page_path)

    return updated


def write_manifest(records):
    fieldnames = [
        "published_date",
        "year",
        "first_author",
        "title",
        "article_url",
        "pdf_url",
        "filename",
        "legacy_filename",
        "local_path",
        "source_page",
        "focus",
    ]
    with MANIFEST_PATH.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, delimiter="\t")
        writer.writeheader()
        for record in records:
            writer.writerow({key: record[key] for key in fieldnames})


def write_xlsx(records):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "papers"

    headers = [
        "No",
        "First Author",
        "Year",
        "Published Date",
        "Title",
        "Filename",
        "Legacy Filename",
        "Local PDF Path",
        "Source Page",
        "Article URL",
        "PDF URL",
        "Focus",
    ]
    sheet.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    for index, record in enumerate(records, start=1):
        sheet.append(
            [
                index,
                record["first_author"],
                record["year"],
                record["published_date"],
                record["title"],
                record["filename"],
                record["legacy_filename"],
                record["local_path"],
                record["source_page"],
                record["article_url"],
                record["pdf_url"],
                record["focus"],
            ]
        )

    for row in range(2, len(records) + 2):
        sheet.cell(row=row, column=10).hyperlink = sheet.cell(row=row, column=10).value
        sheet.cell(row=row, column=11).hyperlink = sheet.cell(row=row, column=11).value

    widths = {
        "A": 6,
        "B": 16,
        "C": 8,
        "D": 16,
        "E": 80,
        "F": 52,
        "G": 64,
        "H": 48,
        "I": 48,
        "J": 46,
        "K": 46,
        "L": 48,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    sheet.freeze_panes = "A2"
    workbook.save(XLSX_PATH)


def append_log_entry(records):
    marker = "corpus normalized | Long Read WGS filenames and spreadsheet"
    existing = LOG_PATH.read_text(encoding="utf-8")
    if marker in existing:
        return

    timestamp = datetime.now().astimezone().strftime("%Y-%m-%d %H:%M %Z")
    lines = [
        "",
        "## [%s] %s" % (timestamp, marker),
        "",
        "- Renamed 10 PDF files to the `first_author_year_first5words.pdf` convention.",
        "- Saved spreadsheet %s." % XLSX_PATH.name,
        "- Updated source page raw file references to the renamed PDFs.",
    ]
    LOG_PATH.write_text(existing.rstrip() + "\n" + "\n".join(lines) + "\n", encoding="utf-8")


def build_parser():
    parser = argparse.ArgumentParser(
        description="Normalize the Long Read WGS starter corpus inside a collection workspace."
    )
    parser.add_argument(
        "--collection",
        default="longread-sequencing",
        help="Collection name under collections/. Defaults to longread-sequencing.",
    )
    parser.add_argument(
        "--workspace",
        help="Explicit workspace path. Overrides --collection.",
    )
    return parser


def main():
    parser = build_parser()
    args = parser.parse_args()
    workspace = resolve_workspace(
        collection=args.collection,
        workspace=args.workspace,
        default_collection="longread-sequencing",
    )
    configure_workspace(workspace.root)

    records = [enrich_record(record) for record in RECORDS]
    renamed = rename_pdfs(records)
    updated_pages = update_source_pages(records)
    write_manifest(records)
    write_xlsx(records)
    append_log_entry(records)

    print("Renamed PDFs: %d" % len(renamed))
    for old_name, new_name in renamed:
        print("- %s -> %s" % (old_name, new_name))

    print("Updated source pages: %d" % len(updated_pages))
    print("Wrote manifest: %s" % MANIFEST_PATH.name)
    print("Wrote spreadsheet: %s" % XLSX_PATH.name)


if __name__ == "__main__":
    main()
