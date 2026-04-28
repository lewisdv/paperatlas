#!/usr/bin/env python3

import argparse
import csv
import re
import unicodedata
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from workspace import PROJECT_ROOT, resolve_workspace


ROOT = PROJECT_ROOT
RAW_DIR = ROOT / "raw" / "sources"
MANIFEST_PATH = ROOT / "organoid_protocols_manifest.tsv"
XLSX_PATH = ROOT / "organoid_protocols_papers.xlsx"
LOG_PATH = ROOT / "wiki" / "log.md"

RECORDS = [
    {
        "published_date": "2011-11-10",
        "first_author": "McCracken",
        "title": "Generating human intestinal tissue from pluripotent stem cells in vitro",
        "article_url": "https://www.nature.com/articles/nprot.2011.410",
        "pdf_url": "https://www.nature.com/articles/nprot.2011.410.pdf",
        "organ": "intestine",
        "focus": "foundational intestinal organoid generation from hPSCs",
    },
    {
        "published_date": "2014-09-04",
        "first_author": "Lancaster",
        "title": "Generation of cerebral organoids from human pluripotent stem cells",
        "article_url": "https://www.nature.com/articles/nprot.2014.158",
        "pdf_url": "https://www.nature.com/articles/nprot.2014.158.pdf",
        "organ": "brain",
        "focus": "foundational cerebral organoid generation from hPSCs",
    },
    {
        "published_date": "2016-01-21",
        "first_author": "Drost",
        "title": "Organoid culture systems for prostate epithelial and cancer tissue",
        "article_url": "https://www.nature.com/articles/nprot.2016.006",
        "pdf_url": "https://www.nature.com/articles/nprot.2016.006.pdf",
        "organ": "prostate",
        "focus": "prostate epithelial and cancer organoid establishment",
    },
    {
        "published_date": "2019-12-18",
        "first_author": "Cattaneo",
        "title": "Tumor organoid-T-cell coculture systems",
        "article_url": "https://www.nature.com/articles/s41596-019-0232-9",
        "pdf_url": "https://www.nature.com/articles/s41596-019-0232-9.pdf",
        "organ": "tumor",
        "focus": "tumor organoid immune coculture workflow",
    },
    {
        "published_date": "2020-09-14",
        "first_author": "Driehuis",
        "title": "Establishment of patient-derived cancer organoids for drug-screening applications",
        "article_url": "https://www.nature.com/articles/s41596-020-0379-4",
        "pdf_url": "https://www.nature.com/articles/s41596-020-0379-4.pdf",
        "organ": "cancer",
        "focus": "patient-derived cancer organoids for screening",
    },
    {
        "published_date": "2020-11-27",
        "first_author": "Hendriks",
        "title": "Establishment of human fetal hepatocyte organoids and CRISPR-Cas9-based gene knockin and knockout in organoid cultures from human liver",
        "article_url": "https://www.nature.com/articles/s41596-020-00411-2",
        "pdf_url": "https://www.nature.com/articles/s41596-020-00411-2.pdf",
        "organ": "liver",
        "focus": "liver organoid establishment plus CRISPR editing",
    },
    {
        "published_date": "2021-01-11",
        "first_author": "Koike",
        "title": "Engineering human hepato-biliary-pancreatic organoids from pluripotent stem cells",
        "article_url": "https://www.nature.com/articles/s41596-020-00441-w",
        "pdf_url": "https://www.nature.com/articles/s41596-020-00441-w.pdf",
        "organ": "hepato-biliary-pancreatic",
        "focus": "multi-lineage hepato-biliary-pancreatic organoid engineering",
    },
    {
        "published_date": "2021-03-10",
        "first_author": "Dekkers",
        "title": "Long-term culture, genetic manipulation and xenotransplantation of human normal and breast cancer organoids",
        "article_url": "https://www.nature.com/articles/s41596-020-00474-1",
        "pdf_url": "https://www.nature.com/articles/s41596-020-00474-1.pdf",
        "organ": "breast",
        "focus": "breast organoid derivation, engineering, and xenotransplantation",
    },
    {
        "published_date": "2021-11-10",
        "first_author": "Drakhlis",
        "title": "Generation of heart-forming organoids from human pluripotent stem cells",
        "article_url": "https://www.nature.com/articles/s41596-021-00629-8",
        "pdf_url": "https://www.nature.com/articles/s41596-021-00629-8.pdf",
        "organ": "heart",
        "focus": "heart-forming organoid generation from hPSCs",
    },
    {
        "published_date": "2023-09-28",
        "first_author": "Vanslambrouck",
        "title": "Generation of proximal tubule-enhanced kidney organoids from human pluripotent stem cells",
        "article_url": "https://www.nature.com/articles/s41596-023-00880-1",
        "pdf_url": "https://www.nature.com/articles/s41596-023-00880-1.pdf",
        "organ": "kidney",
        "focus": "kidney organoid generation with improved proximal tubules",
    },
    {
        "published_date": "2014-11-20",
        "first_author": "Xia",
        "title": "The generation of kidney organoids by differentiation of human pluripotent cells to ureteric bud progenitor-like cells",
        "article_url": "https://www.nature.com/articles/nprot.2014.182",
        "pdf_url": "https://www.nature.com/articles/nprot.2014.182.pdf",
        "organ": "kidney",
        "focus": "ureteric bud progenitor-like route for kidney organoid generation",
    },
    {
        "published_date": "2016-08-18",
        "first_author": "Takasato",
        "title": "Generation of kidney organoids from human pluripotent stem cells",
        "article_url": "https://www.nature.com/articles/nprot.2016.098",
        "pdf_url": "https://www.nature.com/articles/nprot.2016.098.pdf",
        "organ": "kidney",
        "focus": "classic nephron-rich kidney organoid protocol",
    },
    {
        "published_date": "2016-12-22",
        "first_author": "Morizane",
        "title": "Generation of nephron progenitor cells and kidney organoids from human pluripotent stem cells",
        "article_url": "https://www.nature.com/articles/nprot.2016.170",
        "pdf_url": "https://www.nature.com/articles/nprot.2016.170.pdf",
        "organ": "kidney",
        "focus": "high-efficiency nephron progenitor and kidney organoid generation",
    },
    {
        "published_date": "2016-08-25",
        "first_author": "Broutier",
        "title": "Culture and establishment of self-renewing human and mouse adult liver and pancreas 3D organoids and their genetic manipulation",
        "article_url": "https://www.nature.com/articles/nprot.2016.097",
        "pdf_url": "https://www.nature.com/articles/nprot.2016.097.pdf",
        "organ": "liver-pancreas",
        "focus": "adult liver and pancreas organoid culture and genetic manipulation",
    },
    {
        "published_date": "2018-08-23",
        "first_author": "Sloan",
        "title": "Generation and assembly of human brain region-specific three-dimensional cultures",
        "article_url": "https://www.nature.com/articles/s41596-018-0032-7",
        "pdf_url": "https://www.nature.com/articles/s41596-018-0032-7.pdf",
        "organ": "brain",
        "focus": "region-specific brain organoids and assembloids",
    },
    {
        "published_date": "2021-01-14",
        "first_author": "Giandomenico",
        "title": "Generation and long-term culture of advanced cerebral organoids for studying later stages of neural development",
        "article_url": "https://www.nature.com/articles/s41596-020-00433-w",
        "pdf_url": "https://www.nature.com/articles/s41596-020-00433-w.pdf",
        "organ": "brain",
        "focus": "advanced long-term cerebral organoid culture",
    },
    {
        "published_date": "2021-08-11",
        "first_author": "Puschhof",
        "title": "Intestinal organoid cocultures with microbes",
        "article_url": "https://www.nature.com/articles/s41596-021-00589-z",
        "pdf_url": "https://www.nature.com/articles/s41596-021-00589-z.pdf",
        "organ": "colon-intestine",
        "focus": "colon and intestinal organoid microbial coculture workflows",
    },
    {
        "published_date": "2021-10-18",
        "first_author": "Co",
        "title": "Controlling the polarity of human gastrointestinal organoids to investigate epithelial biology and infectious diseases",
        "article_url": "https://www.nature.com/articles/s41596-021-00607-0",
        "pdf_url": "https://www.nature.com/articles/s41596-021-00607-0.pdf",
        "organ": "colon-intestine",
        "focus": "polarity control for gastrointestinal and colon organoid experiments",
    },
    {
        "published_date": "2022-02-02",
        "first_author": "Watanabe",
        "title": "Transplantation of intestinal organoids into a mouse model of colitis",
        "article_url": "https://www.nature.com/articles/s41596-021-00658-3",
        "pdf_url": "https://www.nature.com/articles/s41596-021-00658-3.pdf",
        "organ": "colon-intestine",
        "focus": "intestinal organoid transplantation in colitis repair model",
    },
    {
        "published_date": "2023-05-10",
        "first_author": "MatkovicLeko",
        "title": "A distal lung organoid model to study interstitial lung disease, viral infection and human lung development",
        "article_url": "https://www.nature.com/articles/s41596-023-00827-6",
        "pdf_url": "https://www.nature.com/articles/s41596-023-00827-6.pdf",
        "organ": "lung",
        "focus": "distal lung organoid differentiation and maintenance",
    },
    {
        "published_date": "2018-11-23",
        "first_author": "Broda",
        "title": "Generation of human antral and fundic gastric organoids from pluripotent stem cells",
        "article_url": "https://www.nature.com/articles/s41596-018-0080-z",
        "pdf_url": "https://www.nature.com/articles/s41596-018-0080-z.pdf",
        "organ": "gastric",
        "focus": "gastric antral and fundic organoid generation from hPSCs",
    },
    {
        "published_date": "2019-01-21",
        "first_author": "Miller",
        "title": "Generation of lung organoids from human pluripotent stem cells in vitro",
        "article_url": "https://www.nature.com/articles/s41596-018-0104-8",
        "pdf_url": "https://www.nature.com/articles/s41596-018-0104-8.pdf",
        "organ": "lung",
        "focus": "foundational lung organoid generation from hPSCs",
    },
    {
        "published_date": "2022-03-23",
        "first_author": "Lee",
        "title": "Generation and characterization of hair-bearing skin organoids from human pluripotent stem cells",
        "article_url": "https://www.nature.com/articles/s41596-022-00681-y",
        "pdf_url": "https://www.nature.com/articles/s41596-022-00681-y.pdf",
        "organ": "skin",
        "focus": "hair-bearing skin organoid generation and characterization",
    },
    {
        "published_date": "2024-03-26",
        "first_author": "Olijnik",
        "title": "Generating human bone marrow organoids for disease modeling and drug discovery",
        "article_url": "https://www.nature.com/articles/s41596-024-00971-7",
        "pdf_url": "https://www.nature.com/articles/s41596-024-00971-7.pdf",
        "organ": "bone-marrow",
        "focus": "bone marrow organoid generation for disease modeling and screening",
    },
    {
        "published_date": "2024-05-03",
        "first_author": "Fitzgerald",
        "title": "Generation of ‘semi-guided’ cortical organoids with complex neural oscillations",
        "article_url": "https://www.nature.com/articles/s41596-024-00994-0",
        "pdf_url": "https://www.nature.com/articles/s41596-024-00994-0.pdf",
        "organ": "brain",
        "focus": "semi-guided cortical organoids with electrophysiology workflows",
    },
    {
        "published_date": "2024-07-29",
        "first_author": "Kelley",
        "title": "Host circuit engagement of human cortical organoids transplanted in rodents",
        "article_url": "https://www.nature.com/articles/s41596-024-01029-4",
        "pdf_url": "https://www.nature.com/articles/s41596-024-01029-4.pdf",
        "organ": "brain",
        "focus": "cortical organoid transplantation and host circuit integration",
    },
    {
        "published_date": "2025-06-27",
        "first_author": "Ullah",
        "title": "Generating and characterizing human telencephalic brain organoids from stem cell-derived single neural rosettes",
        "article_url": "https://www.nature.com/articles/s41596-025-01197-x",
        "pdf_url": "https://www.nature.com/articles/s41596-025-01197-x.pdf",
        "organ": "brain",
        "focus": "single-rosette telencephalic brain organoid generation and characterization",
    },
    {
        "published_date": "2025-10-30",
        "first_author": "Dardano",
        "title": "Production of human blood-generating heart-forming organoids and sample preparation for advanced imaging",
        "article_url": "https://www.nature.com/articles/s41596-025-01268-z",
        "pdf_url": "https://www.nature.com/articles/s41596-025-01268-z.pdf",
        "organ": "heart",
        "focus": "blood-generating heart-forming organoids and advanced imaging",
    },
    {
        "published_date": "2025-12-19",
        "first_author": "Meng",
        "title": "CRISPR screens in human neural organoids and assembloids",
        "article_url": "https://www.nature.com/articles/s41596-025-01299-6",
        "pdf_url": "https://www.nature.com/articles/s41596-025-01299-6.pdf",
        "organ": "brain",
        "focus": "CRISPR screening in neural organoids and assembloids",
    },
    {
        "published_date": "2020-06-26",
        "first_author": "Eura",
        "title": "Brainstem Organoids From Human Pluripotent Stem Cells",
        "article_url": "https://www.frontiersin.org/journals/neuroscience/articles/10.3389/fnins.2020.00538/full",
        "pdf_url": "https://www.frontiersin.org/journals/neuroscience/articles/10.3389/fnins.2020.00538/pdf",
        "organ": "brainstem",
        "focus": "brainstem organoid generation with midbrain and hindbrain progenitors",
    },
    {
        "published_date": "2020-07-14",
        "first_author": "Pomeshchik",
        "title": "Human iPSC-Derived Hippocampal Spheroids: An Innovative Tool for Stratifying Alzheimer Disease Patient-Specific Cellular Phenotypes and Developing Therapies",
        "article_url": "https://www.sciencedirect.com/science/article/pii/S2213671120301922",
        "pdf_url": "https://pmc.ncbi.nlm.nih.gov/articles/instance/7363942/bin/mmc6.pdf",
        "organ": "hippocampus",
        "focus": "hippocampal spheroid generation from hiPSCs for disease modeling",
    },
    {
        "published_date": "2021-05-04",
        "first_author": "Zagare",
        "title": "A robust protocol for the generation of human midbrain organoids",
        "article_url": "https://www.sciencedirect.com/science/article/pii/S2666166721002318",
        "pdf_url": "https://pmc.ncbi.nlm.nih.gov/articles/PMC8121770/pdf/main.pdf",
        "organ": "midbrain",
        "focus": "robust human midbrain organoid generation from neuroepithelial stem cells",
    },
    {
        "published_date": "2021-07-08",
        "first_author": "Valiulahi",
        "title": "Generation of caudal-type serotonin neurons and hindbrain-fate organoids from hPSCs",
        "article_url": "https://www.sciencedirect.com/science/article/pii/S221367112100312X",
        "pdf_url": "https://pmc.ncbi.nlm.nih.gov/articles/instance/8365029/bin/mmc2.pdf",
        "organ": "hindbrain",
        "focus": "caudal hindbrain organoid generation with serotonin neuron enrichment",
    },
    {
        "published_date": "2023-07-21",
        "first_author": "Chen",
        "title": "Protocol for generating reproducible miniaturized controlled midbrain organoids",
        "article_url": "https://www.sciencedirect.com/science/article/pii/S2666166723004185",
        "pdf_url": "https://pmc.ncbi.nlm.nih.gov/articles/PMC10382973/pdf/main.pdf",
        "organ": "midbrain",
        "focus": "miniaturized controlled midbrain organoid generation for screening workflows",
    },
    {
        "published_date": "2024-12-02",
        "first_author": "Atamian",
        "title": "Generation and long-term culture of human cerebellar organoids from pluripotent stem cells",
        "article_url": "https://www.nature.com/articles/s41596-024-01093-w",
        "pdf_url": "https://www.nature.com/articles/s41596-024-01093-w.pdf",
        "organ": "cerebellum",
        "focus": "human cerebellar organoid generation and long-term culture",
    },
    {
        "published_date": "2018-10-31",
        "first_author": "Bergmann",
        "title": "Blood–brain-barrier organoids for investigating the permeability of CNS therapeutics",
        "article_url": "https://www.nature.com/articles/s41596-018-0066-x",
        "pdf_url": "https://www.nature.com/articles/s41596-018-0066-x.pdf",
        "organ": "blood-brain barrier",
        "focus": "blood-brain-barrier organoid generation for CNS therapeutic permeability assays",
    },
    {
        "published_date": "2019-05-03",
        "first_author": "Dekkers",
        "title": "High-resolution 3D imaging of fixed and cleared organoids",
        "article_url": "https://www.nature.com/articles/s41596-019-0160-8",
        "pdf_url": "https://www.nature.com/articles/s41596-019-0160-8.pdf",
        "organ": "multi-organ imaging",
        "focus": "whole-organoid clearing and high-resolution 3D imaging workflow",
    },
    {
        "published_date": "2019-05-20",
        "first_author": "Tysoe",
        "title": "Isolation and propagation of primary human cholangiocyte organoids for the generation of bioengineered biliary tissue",
        "article_url": "https://www.nature.com/articles/s41596-019-0168-0",
        "pdf_url": "https://www.nature.com/articles/s41596-019-0168-0.pdf",
        "organ": "biliary",
        "focus": "primary human cholangiocyte organoid derivation and bioengineered biliary tissue generation",
    },
    {
        "published_date": "2020-09-09",
        "first_author": "Sheridan",
        "title": "Establishment and differentiation of long-term trophoblast organoid cultures from the human placenta",
        "article_url": "https://www.nature.com/articles/s41596-020-0381-x",
        "pdf_url": "https://www.nature.com/articles/s41596-020-0381-x.pdf",
        "organ": "placenta",
        "focus": "long-term trophoblast organoid derivation and differentiation from human placenta",
    },
    {
        "published_date": "2022-01-06",
        "first_author": "Miura",
        "title": "Engineering brain assembloids to interrogate human neural circuits",
        "article_url": "https://www.nature.com/articles/s41596-021-00632-z",
        "pdf_url": "https://www.nature.com/articles/s41596-021-00632-z.pdf",
        "organ": "brain assembloid",
        "focus": "brain assembloid engineering for long-range connectivity and circuit interrogation",
    },
    {
        "published_date": "2022-02-18",
        "first_author": "Bakker",
        "title": "Preparing ductal epithelial organoids for high-spatial-resolution molecular profiling using mass spectrometry imaging",
        "article_url": "https://www.nature.com/articles/s41596-021-00661-8",
        "pdf_url": "https://www.nature.com/articles/s41596-021-00661-8.pdf",
        "organ": "ductal epithelial organoids + imaging",
        "focus": "mass spectrometry imaging preparation workflow for ductal epithelial organoids",
    },
    {
        "published_date": "2022-05-11",
        "first_author": "Gurumurthy",
        "title": "Patient-derived and mouse endo-ectocervical organoid generation, genetic manipulation and applications to model infection",
        "article_url": "https://www.nature.com/articles/s41596-022-00695-6",
        "pdf_url": "https://www.nature.com/articles/s41596-022-00695-6.pdf",
        "organ": "cervix",
        "focus": "endo-ectocervical organoid derivation, genetic manipulation, and infection modeling",
    },
    {
        "published_date": "2023-05-17",
        "first_author": "Gabriel",
        "title": "Generation of iPSC-derived human forebrain organoids assembling bilateral eye primordia",
        "article_url": "https://www.nature.com/articles/s41596-023-00814-x",
        "pdf_url": "https://www.nature.com/articles/s41596-023-00814-x.pdf",
        "organ": "forebrain + eye",
        "focus": "forebrain organoids with bilateral eye primordia assembly from iPSCs",
    },
]


def configure_workspace(root: Path):
    global ROOT, RAW_DIR, MANIFEST_PATH, XLSX_PATH, LOG_PATH
    ROOT = root
    RAW_DIR = ROOT / "raw" / "sources"
    MANIFEST_PATH = ROOT / "organoid_protocols_manifest.tsv"
    XLSX_PATH = ROOT / "organoid_protocols_papers.xlsx"
    LOG_PATH = ROOT / "wiki" / "log.md"


def ascii_text(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    return normalized.encode("ascii", "ignore").decode("ascii")


def slug_token(value: str) -> str:
    value = ascii_text(value).lower()
    value = re.sub(r"[^a-z0-9-]+", "_", value)
    value = re.sub(r"_+", "_", value)
    return value.strip("_")


def title_prefix(title: str) -> str:
    normalized_title = re.sub(r"(?<=\d),(?=\d)", "", ascii_text(title))
    words = re.findall(r"[A-Za-z0-9-]+", normalized_title)
    return "_".join(word.lower() for word in words[:5])


def enrich_record(record):
    enriched = dict(record)
    enriched["year"] = record["published_date"][:4]
    enriched["filename"] = "%s_%s_%s.pdf" % (
        slug_token(record["first_author"]),
        enriched["year"],
        title_prefix(record["title"]),
    )
    enriched["local_path"] = "raw/sources/%s" % enriched["filename"]
    enriched["source_page"] = "wiki/sources/%s.md" % Path(enriched["filename"]).stem
    return enriched


def write_manifest(records):
    fieldnames = [
        "published_date",
        "year",
        "first_author",
        "organ",
        "title",
        "article_url",
        "pdf_url",
        "filename",
        "local_path",
        "source_page",
        "focus",
    ]
    with MANIFEST_PATH.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, delimiter="\t")
        writer.writeheader()
        for record in records:
            writer.writerow({key: record[key] for key in fieldnames})


def write_xlsx(records):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "papers"

    headers = [
        "No",
        "First Author",
        "Organ",
        "Year",
        "Published Date",
        "Title",
        "Filename",
        "Local PDF Path",
        "Source Page",
        "Article URL",
        "PDF URL",
        "Focus",
    ]
    sheet.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    for index, record in enumerate(records, start=1):
        sheet.append(
            [
                index,
                record["first_author"],
                record["organ"],
                record["year"],
                record["published_date"],
                record["title"],
                record["filename"],
                record["local_path"],
                record["source_page"],
                record["article_url"],
                record["pdf_url"],
                record["focus"],
            ]
        )

    for row in range(2, len(records) + 2):
        sheet.cell(row=row, column=10).hyperlink = sheet.cell(row=row, column=10).value
        sheet.cell(row=row, column=11).hyperlink = sheet.cell(row=row, column=11).value

    widths = {
        "A": 6,
        "B": 18,
        "C": 18,
        "D": 8,
        "E": 16,
        "F": 90,
        "G": 54,
        "H": 42,
        "I": 42,
        "J": 46,
        "K": 46,
        "L": 54,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    sheet.freeze_panes = "A2"
    workbook.save(XLSX_PATH)


def append_log_entry():
    marker = "corpus curated | Organoid protocol starter corpus"
    existing = LOG_PATH.read_text(encoding="utf-8")
    if marker in existing:
        return

    timestamp = datetime.now().astimezone().strftime("%Y-%m-%d %H:%M %Z")
    lines = [
        "",
        "## [%s] %s" % (timestamp, marker),
        "",
        "- Saved manifest %s." % MANIFEST_PATH.name,
        "- Saved spreadsheet %s." % XLSX_PATH.name,
        "- Prepared an expanded organ-specific organoid protocol set for download and ingest.",
    ]
    LOG_PATH.write_text(existing.rstrip() + "\n" + "\n".join(lines) + "\n", encoding="utf-8")


def build_parser():
    parser = argparse.ArgumentParser(
        description="Prepare the organoid protocol starter corpus manifest and spreadsheet."
    )
    parser.add_argument(
        "--collection",
        default="Organoid",
        help="Collection name under collections/. Defaults to organoid.",
    )
    parser.add_argument(
        "--workspace",
        help="Explicit workspace path. Overrides --collection.",
    )
    return parser


def main():
    parser = build_parser()
    args = parser.parse_args()
    workspace = resolve_workspace(
        collection=args.collection,
        workspace=args.workspace,
        default_collection="Organoid",
    )
    configure_workspace(workspace.root)

    records = [enrich_record(record) for record in RECORDS]
    write_manifest(records)
    write_xlsx(records)
    append_log_entry()

    print("Wrote manifest: %s" % MANIFEST_PATH)
    print("Wrote spreadsheet: %s" % XLSX_PATH)


if __name__ == "__main__":
    main()
